"""Parse JSDA bond-trade CSV (and optionally XLSX) into clean records.

JSDA files typically ship as Shift-JIS (cp932) with title rows above the
header. The parser:

1. Auto-detects encoding (utf-8-sig -> cp932 -> shift_jis -> latin-1).
2. Finds the header row by looking for a date-ish marker.
3. Maps header cells to known fields by alias (order-independent).
4. Coerces numbers (stripping ``%``, ``,``, spaces) and dates.

Produces records with stable field names consumed by :mod:`normalize`.
Column map is documented in ``docs/data_sources.md``.
"""

from __future__ import annotations

import csv
import io
import re
from typing import Any, List, Optional

# field -> accepted header substrings (lower-cased, whitespace-collapsed)
_HEADER_ALIASES: dict[str, list[str]] = {
    "trade_date": ["年月日", "取引日", "営業日", "日付", "取引年月日", "date"],
    "issuer_name": ["銘柄名", "発行体", "発行会社", "発行企業", "企業名", "issuer", "name"],
    "isin": ["isin"],
    "coupon_rate": ["利率", "クーポン", "coupon", "表面利率"],
    "maturity_date": ["償還年月日", "償還日", "償還期日", "満期", "maturity"],
    "high_yield": ["最高利回り", "高値利回り", "高値"],
    "low_yield": ["最低利回り", "安値利回り", "安値"],
    "close_yield": ["終値利回り", "終値", "クローズ", "close"],
    "trade_amount": ["取引金額", "出来高", "取引額", "売買代金"],
}

_DATE_HEADER_MARKERS = ("年月日", "取引日", "営業日", "日付")


def _decode(data) -> str:
    if isinstance(data, str):
        return data
    for enc in ("utf-8-sig", "cp932", "shift_jis", "latin-1"):
        try:
            return data.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
    return data.decode("utf-8", errors="replace")


def _norm_header(cell: str) -> str:
    return re.sub(r"\s+", "", str(cell)).lower()


def _find_header(rows: List[List[str]]) -> tuple[int, List[str]]:
    for i, row in enumerate(rows):
        normed = [_norm_header(c) for c in row]
        if any(any(mk in cell for mk in _DATE_HEADER_MARKERS) for cell in normed):
            return i, normed
    # fall back to first non-empty row
    for i, row in enumerate(rows):
        if any(c.strip() for c in row):
            return i, [_norm_header(c) for c in row]
    return -1, []


def _column_index(headers: List[str]) -> dict[str, int]:
    """Map header cells to known fields.

    Two passes so that ``年月日`` (trade date) is not swallowed by
    ``償還年月日`` (maturity): exact matches first, then substring fallback
    for remaining columns/fields.
    """
    col: dict[str, int] = {}
    claimed: set[int] = set()

    # Pass 1: exact (case/space-insensitive) match.
    for idx, h in enumerate(headers):
        for field, aliases in _HEADER_ALIASES.items():
            if field in col:
                continue
            if any(a.lower() == h for a in aliases):
                col[field] = idx
                claimed.add(idx)
                break
    # Pass 2: substring fallback for unclaimed columns.
    for idx, h in enumerate(headers):
        if idx in claimed:
            continue
        for field, aliases in _HEADER_ALIASES.items():
            if field in col:
                continue
            if any(a.lower() in h for a in aliases):
                col[field] = idx
                claimed.add(idx)
                break
    return col


def _num(cell: str) -> Optional[float]:
    if cell is None:
        return None
    s = str(cell).strip().replace(",", "").replace("%", "").replace("　", "").strip()
    if not s or s in {"-", "－", "―"}:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _date(cell: str) -> Optional[str]:
    if not cell:
        return None
    s = str(cell).strip()
    if not s:
        return None
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y年%m月%d日"):
        try:
            import datetime as _dt

            return _dt.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    m = re.match(r"^\s*(\d{4})\D(\d{1,2})\D(\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return None


def _cell(row: List[str], idx: Optional[int]) -> str:
    if idx is None or idx >= len(row):
        return ""
    return (row[idx] or "").strip()


def parse_csv(data, *, encoding: Optional[str] = None) -> List[dict]:
    """Parse JSDA CSV bytes/text into clean records.

    ``encoding`` overrides auto-detection when known.
    """
    text = data.decode(encoding, errors="replace") if (
        encoding and isinstance(data, (bytes, bytearray))
    ) else _decode(data)
    text = text.lstrip("﻿")

    reader = csv.reader(io.StringIO(text))
    rows = [r for r in reader if any((c or "").strip() for c in r)]
    if not rows:
        return []

    header_idx, headers = _find_header(rows)
    if header_idx < 0:
        return []
    col = _column_index(headers)

    out: List[dict] = []
    for row in rows[header_idx + 1:]:
        if not any((c or "").strip() for c in row):
            continue
        td = _date(_cell(row, col.get("trade_date")))
        if not td:
            continue  # skip totals/title spillover rows
        rec = {
            "trade_date": td,
            "issuer_name": _cell(row, col.get("issuer_name")) or None,
            "isin": _cell(row, col.get("isin")) or None,
            "coupon_rate": _num(_cell(row, col.get("coupon_rate"))),
            "maturity_date": _date(_cell(row, col.get("maturity_date"))),
            "high_yield": _num(_cell(row, col.get("high_yield"))),
            "low_yield": _num(_cell(row, col.get("low_yield"))),
            "close_yield": _num(_cell(row, col.get("close_yield"))),
            "trade_amount": _num(_cell(row, col.get("trade_amount"))),
        }
        out.append(rec)
    return out


def parse_xlsx(data: bytes) -> List[dict]:  # pragma: no cover - optional dep
    """Parse JSDA XLSX. Requires ``openpyxl`` (extras: ``[xlsx]``)."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ImportError(
            "openpyxl is required for JSDA .xlsx parsing; install with "
            "`pip install -e .[xlsx]`"
        ) from exc
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = [[(c.value if c.value is not None else "") for c in row] for row in ws.iter_rows()]
    # Render to CSV text and reuse the CSV parser for consistent column mapping.
    buf = io.StringIO()
    writer = csv.writer(buf)
    for r in rows:
        writer.writerow([("" if v is None else str(v)) for v in r])
    return parse_csv(buf.getvalue())


# ---------------------------------------------------------------------------
# Repo rate (東京レポ・レート / TRR)
# ---------------------------------------------------------------------------
#
# The JSDA publishes the Tokyo Repo Rate per business day across tenors. The
# file is either:
#   * **wide** — one date column and one numeric column per tenor
#     (隔日物 / 1週間物 / 1ヶ月物 / … / 12ヶ月物); or
#   * **long** — a date column, a tenor (期間) column and a rate column.
#
# Both shapes normalize to the same record: ``{as_of_date, tenor, rate}``. The
# tenor is captured verbatim from the source (header text in wide form, cell in
# long form) so no tenor vocabulary is invented here. Rate values are coerced
# via the shared ``_num`` (strips ``%``/``,``, treats ``-`` as missing).

_REPO_DATE_ALIASES: list[str] = ["年月日", "取引日", "営業日", "日付", "date"]
_REPO_TENOR_ALIASES: list[str] = ["期間", "満期", "テナー", "term", "tenor", "期限"]
_REPO_RATE_ALIASES: list[str] = ["レート", "金利", "rate", "レポレート", "%", "東京レポ"]
_REPO_DATE_MARKERS = ("年月日", "取引日", "営業日", "日付")


def _find_repo_header(rows: List[List[str]]) -> tuple[int, List[str]]:
    """Locate the repo-rate header row and return its original (un-normalized)
    cells, so wide-format tenor headers keep their source text."""
    for i, row in enumerate(rows):
        normed = [_norm_header(c) for c in row]
        if any(any(mk in cell for mk in _REPO_DATE_MARKERS) for cell in normed):
            return i, list(row)
    for i, row in enumerate(rows):
        if any((c or "").strip() for c in row):
            return i, list(row)
    return -1, []


def _col_first(headers: List[str], aliases: list[str]) -> Optional[int]:
    """First column index whose normalized header contains any alias."""
    for idx, h in enumerate(headers):
        if any(a.lower() in h for a in aliases):
            return idx
    return None


def _first_numeric_col(
    row: List[str], headers: List[str], skip: set[int]
) -> Optional[int]:
    """First column with a parseable number, excluding ``skip`` indices."""
    for idx in range(min(len(row), len(headers))):
        if idx in skip:
            continue
        if _num(_cell(row, idx)) is not None:
            return idx
    return None


def parse_repo_csv(data, *, encoding: Optional[str] = None) -> List[dict]:
    """Parse JSDA repo-rate (TRR) CSV bytes/text into clean records.

    Each record is ``{"as_of_date", "tenor", "rate"}``. Handles wide (one
    column per tenor) and long (a 期間 column + a rate column) layouts.
    ``encoding`` overrides auto-detection when known.
    """
    text = data.decode(encoding, errors="replace") if (
        encoding and isinstance(data, (bytes, bytearray))
    ) else _decode(data)
    text = text.lstrip("﻿")

    reader = csv.reader(io.StringIO(text))
    rows = [r for r in reader if any((c or "").strip() for c in r)]
    if not rows:
        return []

    header_idx, raw_headers = _find_repo_header(rows)
    if header_idx < 0:
        return []
    norm_headers = [_norm_header(c) for c in raw_headers]

    date_col = _col_first(norm_headers, _REPO_DATE_ALIASES)
    tenor_col = _col_first(norm_headers, _REPO_TENOR_ALIASES)
    rate_col = _col_first(norm_headers, _REPO_RATE_ALIASES)

    out: List[dict] = []
    for row in rows[header_idx + 1:]:
        if not any((c or "").strip() for c in row):
            continue
        d = _date(_cell(row, date_col))
        if not d:
            continue  # skip title/total spillover rows

        if tenor_col is not None:
            # Long layout: one record per row (tenor in a cell).
            tenor = _cell(row, tenor_col)
            rc = rate_col
            if rc is None:
                rc = _first_numeric_col(row, raw_headers, {date_col, tenor_col})
            out.append({"as_of_date": d, "tenor": tenor, "rate": _num(_cell(row, rc))})
        else:
            # Wide layout: one record per numeric column (header = tenor).
            # Only the date column is excluded — ``rate_col`` is a long-layout
            # concept and a tenor header that happens to contain "レート"/"%"
            # must not be dropped here.
            for idx in range(min(len(row), len(raw_headers))):
                if idx == date_col:
                    continue
                val = _num(_cell(row, idx))
                if val is None:
                    continue  # blank tenor for this day -> not published
                tenor = (raw_headers[idx] or "").strip()
                out.append({"as_of_date": d, "tenor": tenor, "rate": val})
    return out


def parse_repo_xlsx(data: bytes) -> List[dict]:  # pragma: no cover - optional dep
    """Parse JSDA repo-rate XLSX. Requires ``openpyxl`` (extras: ``[xlsx]``)."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ImportError(
            "openpyxl is required for JSDA repo-rate .xlsx parsing; "
            "install with `pip install -e .[xlsx]`"
        ) from exc
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = [[(c.value if c.value is not None else "") for c in row] for row in ws.iter_rows()]
    buf = io.StringIO()
    writer = csv.writer(buf)
    for r in rows:
        writer.writerow([("" if v is None else str(v)) for v in r])
    return parse_repo_csv(buf.getvalue())
