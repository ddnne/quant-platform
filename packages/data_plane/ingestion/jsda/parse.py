"""Parse JSDA bond-trade CSV (and optionally XLSX) into clean records.

JSDA files typically ship as Shift-JIS (cp932) with title rows above the
header. The parser:

1. Auto-detects encoding (utf-8-sig -> cp932 -> shift_jis -> latin-1).
2. Finds the header row by looking for a date-ish marker.
3. Maps header cells to known fields by alias (order-independent).
4. Coerces numbers (stripping ``%``, ``,``, spaces) and dates.

Produces records with stable field names consumed by :mod:`normalize`.
Column map is documented in ``docs/data_sources.md``.
"""

from __future__ import annotations

import csv
import io
import re
from typing import Any, List, Optional

# field -> accepted header substrings (lower-cased, whitespace-collapsed)
_HEADER_ALIASES: dict[str, list[str]] = {
    "trade_date": ["年月日", "取引日", "営業日", "日付", "取引年月日", "date"],
    "issuer_name": ["銘柄名", "発行体", "発行会社", "発行企業", "企業名", "issuer", "name"],
    "isin": ["isin"],
    "coupon_rate": ["利率", "クーポン", "coupon", "表面利率"],
    "maturity_date": ["償還年月日", "償還日", "償還期日", "満期", "maturity"],
    "high_yield": ["最高利回り", "高値利回り", "高値"],
    "low_yield": ["最低利回り", "安値利回り", "安値"],
    "close_yield": ["終値利回り", "終値", "クローズ", "close"],
    "trade_amount": ["取引金額", "出来高", "取引額", "売買代金"],
}

_DATE_HEADER_MARKERS = ("年月日", "取引日", "営業日", "日付")


def _decode(data) -> str:
    if isinstance(data, str):
        return data
    for enc in ("utf-8-sig", "cp932", "shift_jis", "latin-1"):
        try:
            return data.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
    return data.decode("utf-8", errors="replace")


def _norm_header(cell: str) -> str:
    return re.sub(r"\s+", "", str(cell)).lower()


def _find_header(rows: List[List[str]]) -> tuple[int, List[str]]:
    for i, row in enumerate(rows):
        normed = [_norm_header(c) for c in row]
        if any(any(mk in cell for mk in _DATE_HEADER_MARKERS) for cell in normed):
            return i, normed
    # fall back to first non-empty row
    for i, row in enumerate(rows):
        if any(c.strip() for c in row):
            return i, [_norm_header(c) for c in row]
    return -1, []


def _column_index(headers: List[str]) -> dict[str, int]:
    """Map header cells to known fields.

    Two passes so that ``年月日`` (trade date) is not swallowed by
    ``償還年月日`` (maturity): exact matches first, then substring fallback
    for remaining columns/fields.
    """
    col: dict[str, int] = {}
    claimed: set[int] = set()

    # Pass 1: exact (case/space-insensitive) match.
    for idx, h in enumerate(headers):
        for field, aliases in _HEADER_ALIASES.items():
            if field in col:
                continue
            if any(a.lower() == h for a in aliases):
                col[field] = idx
                claimed.add(idx)
                break
    # Pass 2: substring fallback for unclaimed columns.
    for idx, h in enumerate(headers):
        if idx in claimed:
            continue
        for field, aliases in _HEADER_ALIASES.items():
            if field in col:
                continue
            if any(a.lower() in h for a in aliases):
                col[field] = idx
                claimed.add(idx)
                break
    return col


def _num(cell: str) -> Optional[float]:
    if cell is None:
        return None
    s = str(cell).strip().replace(",", "").replace("%", "").replace("　", "").strip()
    if not s or s in {"-", "－", "―"}:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _date(cell: str) -> Optional[str]:
    if not cell:
        return None
    s = str(cell).strip()
    if not s:
        return None
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y%m%d", "%Y年%m月%d日"):
        try:
            import datetime as _dt

            return _dt.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    m = re.match(r"^\s*(\d{4})\D(\d{1,2})\D(\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return None


def _cell(row: List[str], idx: Optional[int]) -> str:
    if idx is None or idx >= len(row):
        return ""
    return (row[idx] or "").strip()


def parse_csv(data, *, encoding: Optional[str] = None) -> List[dict]:
    """Parse JSDA CSV bytes/text into clean records.

    ``encoding`` overrides auto-detection when known.
    """
    text = data.decode(encoding, errors="replace") if (
        encoding and isinstance(data, (bytes, bytearray))
    ) else _decode(data)
    text = text.lstrip("﻿")

    reader = csv.reader(io.StringIO(text))
    rows = [r for r in reader if any((c or "").strip() for c in r)]
    if not rows:
        return []

    header_idx, headers = _find_header(rows)
    if header_idx < 0:
        return []
    col = _column_index(headers)

    out: List[dict] = []
    for row in rows[header_idx + 1:]:
        if not any((c or "").strip() for c in row):
            continue
        td = _date(_cell(row, col.get("trade_date")))
        if not td:
            continue  # skip totals/title spillover rows
        rec = {
            "trade_date": td,
            "issuer_name": _cell(row, col.get("issuer_name")) or None,
            "isin": _cell(row, col.get("isin")) or None,
            "coupon_rate": _num(_cell(row, col.get("coupon_rate"))),
            "maturity_date": _date(_cell(row, col.get("maturity_date"))),
            "high_yield": _num(_cell(row, col.get("high_yield"))),
            "low_yield": _num(_cell(row, col.get("low_yield"))),
            "close_yield": _num(_cell(row, col.get("close_yield"))),
            "trade_amount": _num(_cell(row, col.get("trade_amount"))),
        }
        out.append(rec)
    return out


def parse_xlsx(data: bytes) -> List[dict]:  # pragma: no cover - optional dep
    """Parse JSDA XLSX. Requires ``openpyxl`` (extras: ``[xlsx]``)."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ImportError(
            "openpyxl is required for JSDA .xlsx parsing; install with "
            "`pip install -e .[xlsx]`"
        ) from exc
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = [[(c.value if c.value is not None else "") for c in row] for row in ws.iter_rows()]
    # Render to CSV text and reuse the CSV parser for consistent column mapping.
    buf = io.StringIO()
    writer = csv.writer(buf)
    for r in rows:
        writer.writerow([("" if v is None else str(v)) for v in r])
    return parse_csv(buf.getvalue())


# ---------------------------------------------------------------------------
# OTC bond reference prices (公社債店頭売買参考統計値)
# ---------------------------------------------------------------------------

_OTC_REFERENCE_ALIASES: dict[str, list[str]] = {
    "publication_label_date": ["発表日付", "発表日", "publicationdate"],
    "quote_effective_date": ["気配基準日", "基準日", "実効日", "quotedate"],
    "security_code": ["銘柄コード", "証券コード", "isinコード", "isin", "code"],
    "bond_name": ["銘柄名", "債券名", "name"],
    "coupon_rate": ["表面利率", "利率", "coupon"],
    "maturity_date": ["償還期日", "償還年月日", "償還日", "maturity"],
    "average_price": ["平均値単価", "平均単価", "平均値価格"],
    "average_yield": ["平均値複利", "平均利回り", "平均値利回り"],
    "median_price": ["中央値単価", "中央単価", "中央値価格"],
    "median_yield": ["中央値複利", "中央利回り", "中央値利回り"],
    "high_price": ["最高値単価", "最高単価", "最高値価格"],
    "high_yield": ["最高値複利", "最高利回り"],
    "low_price": ["最低値単価", "最低単価", "最低値価格"],
    "low_yield": ["最低値複利", "最低利回り"],
    "individual_investor_flag": [
        "個人向け社債等", "個人向け", "individualinvestor"
    ],
}
_OTC_HEADER_MARKERS = ("銘柄コード", "証券コード", "銘柄名", "債券名")

# The official CSV deliberately ships without a header row; JSDA publishes a
# separate ``csvheaderbaisan.xlsx`` for people opening the file in Excel.  The
# positional layout below follows JSDA's ``baisan_csv.pdf`` data-format sheet.
# The 2015 change increased numeric precision and the 2022 change assigned a
# meaning to 銘柄属性・情報2, but neither moved these columns.
_OTC_POSITIONAL_COLUMNS: dict[str, int] = {
    "publication_label_date": 0,
    "security_code": 2,
    "bond_name": 3,
    "maturity_date": 4,
    "coupon_rate": 5,
    "average_yield": 6,   # 平均値・複利
    "average_price": 7,   # 平均値・単価
    "individual_investor_flag": 12,  # 銘柄属性・情報2 (2022-04-04+)
    "high_price": 15,
    "low_price": 17,
    "high_yield": 21,     # 最高値・複利
    "low_yield": 23,      # 最低値・複利
    "median_yield": 25,   # 中央値・複利
    "median_price": 27,
}
_OTC_POSITIONAL_MIN_COLUMNS = 29


def _otc_header_text(cell: Any) -> str:
    # Source workbooks use parentheses and line breaks to split e.g.
    # ``平均値（単価）``. Removing punctuation makes aliases layout-neutral.
    return re.sub(r"[\s()（）［］\[\]・_%％]", "", str(cell)).lower()


def _find_otc_header(rows: List[List[str]]) -> tuple[int, List[str]]:
    for i, row in enumerate(rows):
        normalized = [_otc_header_text(cell) for cell in row]
        if any(any(marker in cell for marker in _OTC_HEADER_MARKERS) for cell in normalized):
            return i, normalized
    return -1, []


def _otc_columns(headers: List[str]) -> dict[str, int]:
    columns: dict[str, int] = {}
    claimed: set[int] = set()
    normalized_aliases = {
        field: [_otc_header_text(alias) for alias in aliases]
        for field, aliases in _OTC_REFERENCE_ALIASES.items()
    }
    for exact in (True, False):
        for index, header in enumerate(headers):
            if index in claimed:
                continue
            for field, aliases in normalized_aliases.items():
                if field in columns:
                    continue
                match = header in aliases if exact else any(alias in header for alias in aliases)
                if match:
                    columns[field] = index
                    claimed.add(index)
                    break
    return columns


def _looks_like_otc_positional_row(row: List[str]) -> bool:
    """Recognize the governed JSDA headerless layout conservatively."""
    if len(row) < _OTC_POSITIONAL_MIN_COLUMNS:
        return False
    source_date = re.sub(r"\D", "", _cell(row, 0))
    code = _cell(row, _OTC_POSITIONAL_COLUMNS["security_code"])
    name = _cell(row, _OTC_POSITIONAL_COLUMNS["bond_name"])
    return len(source_date) == 8 and source_date.isdigit() and bool(code and name)


def parse_otc_reference_csv(
    data,
    *,
    encoding: Optional[str] = None,
    publication_label_date: Optional[str] = None,
    quote_effective_date: Optional[str] = None,
) -> List[dict]:
    """Parse one official OTC-reference CSV without conflating transactions.

    Archive metadata may supply the publication label and calendar-resolved
    quote date when those fields are not repeated inside each source row.
    Neither date is converted into ``available_at`` here.
    """
    text = data.decode(encoding, errors="replace") if (
        encoding and isinstance(data, (bytes, bytearray))
    ) else _decode(data)
    reader = csv.reader(io.StringIO(text.lstrip("﻿")))
    rows = [row for row in reader if any(str(cell or "").strip() for cell in row)]
    if not rows:
        return []
    header_index, headers = _find_otc_header(rows)
    if header_index < 0:
        if not _looks_like_otc_positional_row(rows[0]):
            return []
        columns = dict(_OTC_POSITIONAL_COLUMNS)
        first_data_index = 0
    else:
        columns = _otc_columns(headers)
        first_data_index = header_index + 1
    out: list[dict] = []
    for row_index, row in enumerate(rows[first_data_index:], start=first_data_index):
        source_row_number = row_index + 1
        code = _cell(row, columns.get("security_code"))
        name = _cell(row, columns.get("bond_name"))
        if not code and not name:
            continue
        label = (
            _date(_cell(row, columns.get("publication_label_date")))
            or _date(publication_label_date or "")
        )
        effective = (
            _date(_cell(row, columns.get("quote_effective_date")))
            or _date(quote_effective_date or "")
        )
        out.append({
            "publication_label_date": label,
            "quote_effective_date": effective,
            "security_code": code,
            "bond_name": name,
            "coupon_rate": _num(_cell(row, columns.get("coupon_rate"))),
            "maturity_date": _date(_cell(row, columns.get("maturity_date"))),
            "average_price": _num(_cell(row, columns.get("average_price"))),
            "average_yield": _num(_cell(row, columns.get("average_yield"))),
            "median_price": _num(_cell(row, columns.get("median_price"))),
            "median_yield": _num(_cell(row, columns.get("median_yield"))),
            "high_price": _num(_cell(row, columns.get("high_price"))),
            "high_yield": _num(_cell(row, columns.get("high_yield"))),
            "low_price": _num(_cell(row, columns.get("low_price"))),
            "low_yield": _num(_cell(row, columns.get("low_yield"))),
            "individual_investor_flag": (
                _cell(row, columns.get("individual_investor_flag")) or None
            ),
            "source_row_number": source_row_number,
        })
    return out


def parse_otc_reference_xlsx(
    data: bytes,
    *,
    publication_label_date: Optional[str] = None,
    quote_effective_date: Optional[str] = None,
) -> List[dict]:  # pragma: no cover - optional dependency exercised in dev
    """Parse an OTC-reference XLSX using the same column contract as CSV."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ImportError(
            "openpyxl is required for JSDA OTC-reference .xlsx parsing; "
            "install with `pip install -e .[xlsx]`"
        ) from exc
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    worksheet = workbook.active
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    for row in worksheet.iter_rows(values_only=True):
        writer.writerow(["" if value is None else str(value) for value in row])
    return parse_otc_reference_csv(
        buffer.getvalue(),
        publication_label_date=publication_label_date,
        quote_effective_date=quote_effective_date,
    )


# ---------------------------------------------------------------------------
# Repo rate (東京レポ・レート / TRR)
# ---------------------------------------------------------------------------
#
# The JSDA publishes the Tokyo Repo Rate per business day across tenors. The
# file is either:
#   * **wide** — one date column and one numeric column per tenor
#     (隔日物 / 1週間物 / 1ヶ月物 / … / 12ヶ月物); or
#   * **long** — a date column, a tenor (期間) column and a rate column.
#
# Both shapes normalize to the same record: ``{as_of_date, tenor, rate}``. The
# tenor is captured verbatim from the source (header text in wide form, cell in
# long form) so no tenor vocabulary is invented here. Rate values are coerced
# via the shared ``_num`` (strips ``%``/``,``, treats ``-`` as missing).

_REPO_DATE_ALIASES: list[str] = [
    "年月日", "取引日", "営業日", "公表日", "基準日", "日付", "date"
]
_REPO_TENOR_ALIASES: list[str] = ["期間", "満期", "テナー", "term", "tenor", "期限"]
_REPO_RATE_ALIASES: list[str] = ["レート", "金利", "rate", "レポレート", "%", "東京レポ"]
_REPO_DATE_MARKERS = ("年月日", "取引日", "営業日", "公表日", "基準日", "日付")


def _find_repo_header(rows: List[List[str]]) -> tuple[int, List[str]]:
    """Locate the repo-rate header row and return its original (un-normalized)
    cells, so wide-format tenor headers keep their source text.

    Official ``trrts.xls`` uses a matrix header like
    ``term/ターム | overnight/翌日物 | 1W | ...`` (no explicit date column
    name). Prefer that over the workbook title / footnote rows.
    """
    # 1) Official TRR matrix header (must beat footnotes that mention ターム物/1W).
    for i, row in enumerate(rows):
        first = (row[0] if row else "") or ""
        if first.strip().startswith("※"):
            continue
        normed = [_norm_header(c) for c in row]
        has_overnight = any(
            ("overnight" in c) or ("翌日" in c) for c in normed
        )
        has_term_hdr = any(
            c.startswith("term")
            or c.startswith("tenor")
            or c.startswith("ターム")
            or "term/" in c
            or "tenor/" in c
            for c in normed
        )
        if has_term_hdr and has_overnight:
            return i, list(row)
    # 2) Explicit date-column headers (long/CSV layouts).
    for i, row in enumerate(rows):
        first = (row[0] if row else "") or ""
        if first.strip().startswith("※"):
            continue
        normed = [_norm_header(c) for c in row]
        if any(any(mk in cell for mk in _REPO_DATE_MARKERS) for cell in normed):
            return i, list(row)
    # 3) Fallback: first non-empty non-footnote row.
    for i, row in enumerate(rows):
        first = (row[0] if row else "") or ""
        if first.strip().startswith("※"):
            continue
        if any((c or "").strip() for c in row):
            return i, list(row)
    return -1, []


def _col_first(headers: List[str], aliases: list[str]) -> Optional[int]:
    """First column index whose normalized header contains any alias."""
    for idx, h in enumerate(headers):
        if any(a.lower() in h for a in aliases):
            return idx
    return None


def _first_numeric_col(
    row: List[str], headers: List[str], skip: set[int]
) -> Optional[int]:
    """First column with a parseable number, excluding ``skip`` indices."""
    for idx in range(min(len(row), len(headers))):
        if idx in skip:
            continue
        if _num(_cell(row, idx)) is not None:
            return idx
    return None


def parse_repo_csv(data, *, encoding: Optional[str] = None) -> List[dict]:
    """Parse JSDA repo-rate (TRR) CSV bytes/text into clean records.

    Each record is ``{"as_of_date", "tenor", "rate"}``. Handles wide (one
    column per tenor) and long (a 期間 column + a rate column) layouts.
    ``encoding`` overrides auto-detection when known.
    """
    text = data.decode(encoding, errors="replace") if (
        encoding and isinstance(data, (bytes, bytearray))
    ) else _decode(data)
    text = text.lstrip("﻿")

    reader = csv.reader(io.StringIO(text))
    rows = [r for r in reader if any((c or "").strip() for c in r)]
    if not rows:
        return []

    header_idx, raw_headers = _find_repo_header(rows)
    if header_idx < 0:
        return []
    norm_headers = [_norm_header(c) for c in raw_headers]

    date_col = _col_first(norm_headers, _REPO_DATE_ALIASES)
    tenor_col = _col_first(norm_headers, _REPO_TENOR_ALIASES)
    rate_col = _col_first(norm_headers, _REPO_RATE_ALIASES)

    # Official wide TRR matrix: col0 header is "term/ターム" but cells are dates;
    # other columns are tenors (overnight, 1W, …). Force wide layout.
    wide_matrix = False
    if tenor_col == 0:
        other = [
            (raw_headers[i] or "").strip()
            for i in range(1, len(raw_headers))
            if (raw_headers[i] or "").strip()
        ]
        if other and any(
            re.search(r"(overnight|翌日|1\s*w|1\s*m|2\s*w|3\s*m|/)", h, re.I)
            for h in other
        ):
            tenor_col = None
            date_col = 0
            wide_matrix = True
    if date_col is None:
        for probe in rows[header_idx + 1 : header_idx + 12]:
            if _date(_cell(probe, 0)):
                date_col = 0
                wide_matrix = True
                break

    # Dual header row (official trrts.xls): parent tenors + settlement (T+0/T+1).
    tenor_labels = list(raw_headers)
    data_start = header_idx + 1
    if wide_matrix and header_idx + 1 < len(rows):
        sub = rows[header_idx + 1]
        # If next row looks like settlement labels (T+0/T+1) not data dates.
        sub_joined = " ".join((c or "") for c in sub).upper()
        if "T+0" in sub_joined or "T+1" in sub_joined:
            parent = ""
            labels: list[str] = []
            width = max(len(raw_headers), len(sub))
            for i in range(width):
                top = (raw_headers[i] if i < len(raw_headers) else "") or ""
                bot = (sub[i] if i < len(sub) else "") or ""
                top_s, bot_s = top.strip(), bot.strip()
                if top_s:
                    parent = top_s
                if i == (date_col or 0):
                    labels.append(top_s or "date")
                    continue
                if top_s and bot_s and re.match(r"T\+\d+", bot_s, re.I):
                    labels.append(f"{top_s}/{bot_s}")
                elif bot_s and re.match(r"T\+\d+", bot_s, re.I) and parent:
                    labels.append(f"{parent}/{bot_s}")
                elif top_s:
                    labels.append(top_s)
                elif bot_s:
                    labels.append(bot_s)
                else:
                    labels.append("")
            tenor_labels = labels
            data_start = header_idx + 2

    def _clean_tenor(label: str) -> str:
        t = (label or "").strip()
        if not t or t in {"(%)", "%", "-", "—", "－"}:
            return ""
        # Drop pure rate/unit junk headers.
        if re.fullmatch(r"[\(%）%\s]+", t):
            return ""
        return t

    out: List[dict] = []
    for row in rows[data_start:]:
        if not any((c or "").strip() for c in row):
            continue
        d = _date(_cell(row, date_col))
        if not d:
            continue  # skip title/total spillover rows

        if tenor_col is not None:
            # Long layout: one record per row (tenor in a cell).
            tenor = _cell(row, tenor_col).strip()
            rc = rate_col
            if rc is None:
                rc = _first_numeric_col(row, raw_headers, {date_col, tenor_col})
            rate = _num(_cell(row, rc))
            if not tenor or rate is None:
                continue
            out.append({"as_of_date": d, "tenor": tenor, "rate": rate})
        else:
            # Wide layout: one record per numeric column (header = tenor).
            width = max(len(row), len(tenor_labels))
            for idx in range(width):
                if date_col is not None and idx == date_col:
                    continue
                val = _num(_cell(row, idx))
                if val is None:
                    continue  # blank tenor for this day -> not published
                tenor = _clean_tenor(
                    tenor_labels[idx] if idx < len(tenor_labels) else ""
                )
                if not tenor:
                    continue  # never emit empty tenor (causes duplicate keys)
                out.append({"as_of_date": d, "tenor": tenor, "rate": val})
    return out


def parse_repo_xlsx(data: bytes) -> List[dict]:  # pragma: no cover - optional dep
    """Parse JSDA repo-rate XLSX. Requires ``openpyxl`` (extras: ``[xlsx]``)."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ImportError(
            "openpyxl is required for JSDA repo-rate .xlsx parsing; "
            "install with `pip install -e .[xlsx]`"
        ) from exc
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = [[(c.value if c.value is not None else "") for c in row] for row in ws.iter_rows()]
    buf = io.StringIO()
    writer = csv.writer(buf)
    for r in rows:
        writer.writerow([("" if v is None else str(v)) for v in r])
    return parse_repo_csv(buf.getvalue())


def parse_repo_xls(data: bytes) -> List[dict]:
    """Parse the authoritative legacy ``trrts.xls`` time-series workbook.

    JSDA intentionally publishes the complete Tokyo Repo Rate history as an
    OLE2 ``.xls`` workbook.  ``xlrd`` is therefore a governed dependency, not
    a format that callers may silently skip.  Every sheet is considered and
    the sheet yielding the most canonical observations is selected; Excel
    serial dates are converted using the workbook's own date mode.
    """
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - dependency contract
        raise ImportError(
            "xlrd is required for the authoritative JSDA Tokyo Repo Rate "
            "history; install with `pip install -e .[xls]`"
        ) from exc

    workbook = xlrd.open_workbook(file_contents=data, on_demand=True)
    candidates: list[list[dict]] = []
    try:
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            buffer = io.StringIO()
            writer = csv.writer(buffer)
            for row_index in range(sheet.nrows):
                rendered: list[str] = []
                for column_index in range(sheet.ncols):
                    cell = sheet.cell(row_index, column_index)
                    value = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        value = xlrd.xldate_as_datetime(
                            value, workbook.datemode
                        ).strftime("%Y-%m-%d")
                    elif cell.ctype == xlrd.XL_CELL_NUMBER and isinstance(
                        value, (int, float)
                    ):
                        # Official trrts.xls stores as_of dates as unformatted
                        # Excel serial numbers (e.g. 41211) while rates stay in
                        # [0, 10]. Convert serial-range numbers to YYYY-MM-DD.
                        num = float(value)
                        if 20000.0 <= num <= 60000.0:
                            try:
                                value = xlrd.xldate_as_datetime(
                                    num, workbook.datemode
                                ).strftime("%Y-%m-%d")
                            except (ValueError, OverflowError):
                                value = int(num) if num.is_integer() else num
                        elif num.is_integer():
                            value = int(num)
                    rendered.append("" if value is None else str(value))
                writer.writerow(rendered)
            records = parse_repo_csv(buffer.getvalue())
            if records:
                candidates.append(records)
    finally:
        release = getattr(workbook, "release_resources", None)
        if callable(release):
            release()
    return max(candidates, key=len, default=[])
